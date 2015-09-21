import os
import sys
import xml.etree.ElementTree as ET
import warnings
import easygui as e
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Side
from openpyxl.styles import borders
from openpyxl.cell import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.writer.write_only import WriteOnlyCell
from openpyxl.styles import Style, Font

warnings.filterwarnings("ignore", category=UserWarning)


def create_formula(inicio, fin, fila):
    return '=SUM('+get_column_letter(inicio)+str(fila)+':'+get_column_letter(fin)+str(fila)+')'


def fill_cell(working_sheet, value, negrita=False):
    try:
        cell = WriteOnlyCell(working_sheet, value=value)
        cell.style = Style(font=Font(name='Calibri', size=11, bold=negrita),
                           border=Border(left=Side(border_style=borders.BORDER_THIN,color='FF000000'),
                           right=Side(border_style=borders.BORDER_THIN,color='FF000000'),
                           top=Side(border_style=borders.BORDER_THIN, color='FF000000'),
                           bottom=Side(border_style=borders.BORDER_THIN,color='FF000000')))
        cell.alignment = Alignment(horizontal='center', vertical='center')
    except:
        pass
    return cell


def create_graph(working_sheet, n_operario, n_dias):
    data = Reference(working_sheet, min_col=(n_dias+3), min_row=2, max_row=n_operario, max_col=(n_dias+3))
    cats = Reference(working_sheet, min_col=1, min_row=3, max_row=n_operario)
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 12
    chart2.grouping = "stacked"
    chart2.title = 'Inasistencias por operario en Agosto'
    chart2.y_axis.title = 'Inasistencias'
    chart2.x_axis.title = 'Operario ID'
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    working_sheet.add_chart(chart2, anchor=get_column_letter(6+n_dias)+str(2))


def transform(file, simple):
    # Parse del xml
    tree = ET.parse(file)
    root = tree.getroot()
    tipo = root[0][0][0][0][0].attrib.get('name')


    wb = Workbook()
    ws = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(ws)

    # ws.column_dimensions.group('A', 'D', hidden=True)
    '''
    for ws_column in range(1,10):
        ws.column_dimensions.
        ws.column_dimensions.group('A', 'I', hidden=True)
        col_letter = get_column_letter(ws_column)
        ws.column_dimensions[col_letter].bestFit = True
    '''
    if tipo == "datos_frx2xml":
        ws = wb.create_sheet(title="Personal")
        # Por persona

        first = True

        for f in root.iter('datos_frx2xml'):
            if first:
                first = False
                to_fila = []
                fecha = f.find('f1').text

                if fecha is not None:
                    to_fila.append(fill_cell(ws, "Fecha Emision: ", True))
                    to_fila.append(fill_cell(ws, fecha))

                periodo = f.find('f3').text

                if periodo is not None:
                    to_fila.append(fill_cell(ws, "Periodo: ", True))
                    to_fila.append(fill_cell(ws, periodo))

                tipo = f.find('f4').text

                if tipo is not None:
                    tipo = tipo.split(':')
                    to_fila.append(fill_cell(ws, tipo[1]))

                departamento = f.find('f5').text

                if departamento is not None:
                    departamento = departamento.split(':')
                    to_fila.append(fill_cell(ws, departamento[1]))

                datos = f.find('f6').text

                if datos is not None:
                    datos = datos.split('-')

                    to_fila.append(fill_cell(ws, "Cedula:", True))
                    to_fila.append(fill_cell(ws, int(datos[0])))

                    to_fila.append(fill_cell(ws, "Nombre:", True))
                    to_fila.append(fill_cell(ws, datos[1]))

                ws.append(to_fila)

                to_fila = []

                # Formato de las filas
                to_fila.append(fill_cell(ws, "Fecha ", True))
                to_fila.append(fill_cell(ws, "1er Turno Entrada Calculo ", True))
                to_fila.append(fill_cell(ws, "1er Turno Salida Calculo ", True))
                to_fila.append(fill_cell(ws, "2do Turno Entrada Calculo ", True))
                to_fila.append(fill_cell(ws, "2do Turno Salida Calculo ", True))
                to_fila.append(fill_cell(ws, "1er Turno Entrada Real ", True))
                to_fila.append(fill_cell(ws, "1er Turno Salida Real ", True))
                to_fila.append(fill_cell(ws, "2do Turno Entrada Real ", True))
                to_fila.append(fill_cell(ws, "2do Turno Salida Real ", True))
                to_fila.append(fill_cell(ws, "Horas Trabajadas", True))
                to_fila.append(fill_cell(ws, "Deduccion", True))
                to_fila.append(fill_cell(ws, "No trabajo", True))
                to_fila.append(fill_cell(ws, "Observacion", True))

                ws.append(to_fila)

            to_fila = []
            to_fila.append(fill_cell(ws, f.find('f7').text))

            libre = f.find('f22').text

            if libre is not "LIBRE":
                # Anado lo basico
                to_fila.append(fill_cell(ws, "8:00"))
                to_fila.append(fill_cell(ws, "12:00"))
                to_fila.append(fill_cell(ws, "13:00"))
                to_fila.append(fill_cell(ws, "16:15"))

                # Horas normales
                to_fila.append(fill_cell(ws, f.find('f12').text))
                to_fila.append(fill_cell(ws, f.find('f13').text))
                to_fila.append(fill_cell(ws, f.find('f14').text))
                to_fila.append(fill_cell(ws, f.find('f15').text))

                # Total de horas trabajadas
                to_fila.append(fill_cell(ws, f.find('f20').text))

                # Deduccion
                to_fila.append(fill_cell(ws, f.find('f18').text))

                # No Trabajo
                to_fila.append(fill_cell(ws, f.find('f19').text))

            # Observacion
            to_fila.append(fill_cell(ws, f.find('f22').text))

            ws.append(to_fila)

        wb.save(simple[:-4]+'.xlsx')
    else:
        # Por Grupo
        ws = wb.create_sheet(title="Grupal")

        # arreglo de fechas
        fechas = ['f'+str(i) for i in list(range(4, 19))]+['f'+str(i) for i in list(range(20, 36))]

        # arreglo de inasistencias
        inasistencias = ['f'+str(i) for i in list(range(38, 53))]+['f'+str(i) for i in list(range(54, 70))]

        first = True
        control = [0, 0, 0, 2]

        for f in root.iter('repinas_frx2xml'):
            control[2] = 0
            control[3] += 1
            # Formatos para la primera vez
            if first:
                first = False
                # Primera fila
                col = []

                # Fecha emision reporte
                fecha = f.find('f1').text

                if fecha is not None:
                    col.append('Fecha Emision: ')
                    col.append(fill_cell(ws, fecha, True))

                # Tipo de personal
                personal = f.find('f2').text

                if personal is not None:
                    col.append(fill_cell(ws, personal))

                departamento = f.find('f19').text

                if departamento is not None:
                    col.append(fill_cell(ws, departamento))

                # Agregamos la primera fila al trabajo
                ws.append(col)

                # Formato : id | personal | fechas | Total

                col = []

                # Agregamos nombre y ID
                col.append(fill_cell(ws, 'ID', True))
                col.append(fill_cell(ws, 'Nombre', True))

                # Agregamos las fechas
                for fecha in fechas:
                    nueva = f.find(fecha).text
                    if nueva is not None:
                        col.append(fill_cell(ws, nueva, True))
                    else:
                        control[0] += 1
                    control[1] += 1

                col.append(fill_cell(ws, 'Total', True))
                col.append(fill_cell(ws, 'Total Sistema', True))


                ws.append(col)

            col = []

            # id del obrero
            id = f.find('f36').text
            col.append(fill_cell(ws, int(id)))

            # nombre del obrero
            nombre = f.find('f37').text
            col.append(fill_cell(ws, nombre))

            # Constante de numero de columnas
            n_dias = control[1] - control[0]

            # Agregamos inasistencias y cambiamos formato
            for inasistencia in inasistencias:
                lista = f.find(inasistencia).text
                try:
                    if ord(lista) == 189:
                        lista = 0.5
                    elif ord(lista) == 120:
                        lista = 1
                except:
                    pass
                finally:
                    control[2] += 1
                    if n_dias >= control[2]:
                        col.append(fill_cell(ws, lista))

            # Agregamos la formula para la columna desde 3 hasta n_dias+2
            col.append(fill_cell(ws, create_formula(3, n_dias+2, control[3])))

            # Agregamos el total
            total = f.find('f53').text
            col.append(fill_cell(ws, int(total)))

            ws.append(col)

        # Guardamos el excel
        create_graph(ws, control[3], n_dias)

        wb.save(simple[:-4]+'.xlsx')


def salida(opcion):
    if opcion is not None:
        pass
    else:
        e.msgbox("Cerrando aplicacion")
        sys.exit(0)

yes = True

while yes:
    mensaje = e.msgbox("Transformacion de archivos xml"+"\n"+"\tVersion 1.0.0"+"\n", image='images\\LogoRonava.png')
    salida(mensaje)

    directorio = e.diropenbox(title = "\t\tEscoger directorio", msg = "\n"+"\tSeleccione el directorio con los archivos csv")
    salida(directorio)

    opciones = next(os.walk(directorio))[2]
    salida(opciones)

    # Seleccionamos solo las opciones .csv

    xml = list()

    for file in opciones:
        if file[-4:] == ".xml":
            xml.append(file)

    archivos = e.multchoicebox(msg='Seleccione los archivos a transformar', title='Seleccion de archivos', choices=xml)
    salida(archivos)

    salida(e.msgbox("Iniciar transformacion"))

    for file in archivos:
        try:
            transform((directorio+"%s"+file) % '\\\\', file)
        except IOError:
            msg = "El archivo "+file+" esta abierto en Excel."+"\n"+"Cierrelo para poder transformarlo"
            title = "Error transformando"

            opcion = e.ccbox(msg, title)
            salida(opcion)

    msg = "Reiniciar proceso?"
    title = "Por favor seleccione una opcion"
    if e.ccbox(msg, title):     # show a Continue/Cancel dialog
        pass  # user chose Continue
    else:
        sys.exit(0)     # user chose Cancel



