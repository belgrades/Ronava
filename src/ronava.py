import os
import sys
import xml.etree.ElementTree as ET
import warnings
import easygui as e
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Side
from openpyxl.styles import borders
from openpyxl.cell import get_column_letter
from openpyxl.chart import BarChart, Reference, Series, LineChart
from openpyxl.writer.write_only import WriteOnlyCell
from openpyxl.styles import Style, Font
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning)

# TODO cambiar ejes en grafico ID

calendario = dict()

def create_formula(col_ini, col_fin, row_ini, row_fin, days):
    # TODO modify name to be consistent
    return '='+str(days)+'-SUM('+get_column_letter(col_ini)+str(row_ini)+':'+get_column_letter(col_fin)+str(row_fin)+')'


def fill_cell(working_sheet, value, negrita=False):
    # TODO add colors depending on the day (labor or not)
    try:
        cell = WriteOnlyCell(working_sheet, value=value)
        cell.style = Style(font=Font(name='Calibri', size=11, bold=negrita),
                           border=Border(left=Side(border_style=borders.BORDER_THIN,color='FF000000'),
                           right=Side(border_style=borders.BORDER_THIN,color='FF000000'),
                           top=Side(border_style=borders.BORDER_THIN, color='FF000000'),
                           bottom=Side(border_style=borders.BORDER_THIN,color='FF000000')))
        cell.alignment = Alignment(horizontal='center', vertical='center')
    except:
        pass
    return cell


def ronava_bar_chart(writingSheet, dataSheet, params):
    # TODO add dictionary in parameters to avoid overlapping
    if params["use"] == "bars":
        data = Reference(dataSheet,
                         min_col=params['data_min_col'],
                         min_row=params['data_min_row'],
                         max_row=params['data_max_row'],
                         max_col=params['data_max_col'])
        cats = Reference(dataSheet,
                         min_col=params['cats_min_col'],
                         min_row=params['cats_min_row'],
                         max_row=params['cats_max_row'],
                         max_col=params['cats_max_col'])
        chart = BarChart()
        chart.type = params['type']
        chart.style = 12
        # chart.grouping = "stacked"
        chart.title = params['title']
        chart.y_axis.title = params['y_axis']
        chart.x_axis.title = params['x_axis']
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = params['heigth']
        chart.width = params['width']
        writingSheet.add_chart(chart, 'D2')
    elif params["use"] == 'single':
        c1 = BarChart()
        v1 = Reference(dataSheet,
                       min_col=params['data_min_col'],
                       min_row=params['data_min_row'],
                       max_col=params['data_max_col'])

        cats = Reference(dataSheet,
                        min_col=params['cats_min_col'],
                        min_row=params['cats_min_row'],
                        max_col=params['cats_max_col'])
        c1.series = [Series(v1, title_from_data=True)]
        c1.style = 12
        c1.set_categories(cats)
        c1.x_axis.title = params['x_axis']
        c1.y_axis.title = params['y_axis']
        c1.height = params['heigth']
        c1.width = params['width']
        c1.title = params['title']
        writingSheet.add_chart(c1, "D4")
    else:
        c1 = BarChart()
        v1 = Reference(dataSheet,
                       min_col=params['data_min_col'],
                       min_row=params['data_min_row'],
                       max_col=params['data_max_col'])

        cats = Reference(dataSheet,
                        min_col=params['cats_min_col'],
                        min_row=params['cats_min_row'],
                        max_col=params['cats_max_col'])
        c1.series = [Series(v1, title_from_data=True)]
        c1.y_axis.majorGridlines = None
        c1.set_categories(cats)
        c1.x_axis.title = params['x_axis']
        c1.y_axis.title = params['y_axis']
        c1.height = params['heigth']
        c1.width = params['width']
        c1.title = params['title']
        c1.style = 12
        # Create a second chart
        c2 = LineChart()
        v2 = Reference(dataSheet,
                       min_col=params['data_min_col'],
                       min_row=params['data_min_row']+1,
                       max_col=params['data_max_col'])
        c2.series = [Series(v2, title_from_data=True)]
        c2.y_axis.axId = 20
        c2.y_axis.title = "Porcentaje Produccion"
        # Assign the y-axis of the second chart to the third axis of the first chart
        c1.z_axis = c2.y_axis
        c1.y_axis.crosses = "max"
        c1 += c2

        writingSheet.add_chart(c1, "D4")


def transform(file, simple, directorio):
    # TODO add more comments to remember
    # TODO change size of charts
    # TODO add more charts
    # TODO change non assistance to assistance
    # TODO change control to dictionary
    # TODO modify columns of hidden
    global calendario
    # Parse del xml
    tree = ET.parse(file)
    root = tree.getroot()
    tipo = root[0][0][0][0][0].attrib.get('name')

    wb = Workbook()
    ws = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(ws)

    # ws.column_dimensions.group('A', 'D', hidden=True)

    if tipo == "datos_frx2xml":
        ws = wb.create_sheet(title="Personal")
        # Por persona

        first = True
        month = ''

        for f in root.iter('datos_frx2xml'):
            if first:
                first = False
                to_fila = []
                fecha = f.find('f1').text

                if fecha is not None:
                    to_fila.append(fill_cell(ws, "Fecha Emision: ", True))
                    to_fila.append(fill_cell(ws, fecha))

                periodo = f.find('f3').text

                if periodo is not None:
                    to_fila.append(fill_cell(ws, "Periodo: ", True))
                    to_fila.append(fill_cell(ws, periodo))

                tipo = f.find('f4').text

                if tipo is not None:
                    tipo = tipo.split(':')
                    to_fila.append(fill_cell(ws, tipo[1]))

                departamento = f.find('f5').text

                if departamento is not None:
                    departamento = departamento.split(':')
                    to_fila.append(fill_cell(ws, departamento[1]))

                datos = f.find('f6').text

                if datos is not None:
                    datos = datos.split('-')

                    to_fila.append(fill_cell(ws, "Cedula:", True))
                    to_fila.append(fill_cell(ws, int(datos[0])))

                    to_fila.append(fill_cell(ws, "Nombre:", True))
                    to_fila.append(fill_cell(ws, datos[1]))

                ws.append(to_fila)

                to_fila = []

                # Formato de las filas
                to_fila.append(fill_cell(ws, "Fecha ", True))
                to_fila.append(fill_cell(ws, "1er Turno Entrada Calculo ", True))
                to_fila.append(fill_cell(ws, "1er Turno Salida Calculo ", True))
                to_fila.append(fill_cell(ws, "2do Turno Entrada Calculo ", True))
                to_fila.append(fill_cell(ws, "2do Turno Salida Calculo ", True))
                to_fila.append(fill_cell(ws, "1er Turno Entrada Real ", True))
                to_fila.append(fill_cell(ws, "1er Turno Salida Real ", True))
                to_fila.append(fill_cell(ws, "2do Turno Entrada Real ", True))
                to_fila.append(fill_cell(ws, "2do Turno Salida Real ", True))
                to_fila.append(fill_cell(ws, "Horas Trabajadas", True))
                to_fila.append(fill_cell(ws, "Deduccion", True))
                to_fila.append(fill_cell(ws, "No trabajo", True))
                to_fila.append(fill_cell(ws, "Observacion", True))

                ws.append(to_fila)

            to_fila = []
            to_fila.append(fill_cell(ws, f.find('f7').text))

            libre = f.find('f22').text

            if libre is not "LIBRE":
                # Anado lo basico
                to_fila.append(fill_cell(ws, "8:00"))
                to_fila.append(fill_cell(ws, "12:00"))
                to_fila.append(fill_cell(ws, "13:00"))
                to_fila.append(fill_cell(ws, "16:15"))

                # Horas normales
                to_fila.append(fill_cell(ws, f.find('f12').text))
                to_fila.append(fill_cell(ws, f.find('f13').text))
                to_fila.append(fill_cell(ws, f.find('f14').text))
                to_fila.append(fill_cell(ws, f.find('f15').text))

                # Total de horas trabajadas
                to_fila.append(fill_cell(ws, f.find('f20').text))

                # Deduccion
                to_fila.append(fill_cell(ws, f.find('f18').text))

                # No Trabajo
                to_fila.append(fill_cell(ws, f.find('f19').text))

            # Observacion
            to_fila.append(fill_cell(ws, f.find('f22').text))

            ws.append(to_fila)
    else:
        # Por Grupo
        ws = wb.create_sheet(title="Grupal")

        # Mes de estudio
        month = ''

        # arreglo de fechas
        fechas = ['f'+str(i) for i in list(range(4, 19))]+['f'+str(i) for i in list(range(20, 36))]

        # arreglo de inasistencias
        inasistencias = ['f'+str(i) for i in list(range(38, 53))]+['f'+str(i) for i in list(range(54, 70))]

        first = True
        control = [0, 0, 0, 2]

        for f in root.iter('repinas_frx2xml'):
            control[2] = 0
            control[3] += 1
            # Formatos para la primera vez
            if first:
                first = False
                # Primera fila
                col = []

                # Fecha emision reporte
                fecha = f.find('f1').text

                if fecha is not None:
                    col.append('Fecha Emision: ')
                    col.append(fill_cell(ws, fecha, True))

                # Tipo de personal
                personal = f.find('f2').text

                if personal is not None:
                    col.append(fill_cell(ws, personal))

                departamento = f.find('f19').text

                if departamento is not None:
                    col.append(fill_cell(ws, departamento))

                # Agregamos la primera fila al trabajo
                ws.append(col)

                # Formato : id | personal | fechas | Total

                col = []

                # Agregamos nombre y ID
                col.append(fill_cell(ws, 'ID', True))
                col.append(fill_cell(ws, 'Nombre', True))

                # Agregamos las fechas
                actualDates = []
                for fecha in fechas:
                    nueva = f.find(fecha).text
                    if nueva is not None:
                        nueva = nueva.split('\n')
                        both = nueva[0].split('/')
                        actualDates.append(both[1]+nueva[1]+'/'+both[0])
                        month = both[1]+nueva[1]
                        col.append(fill_cell(ws, both[1]+nueva[1]+'/'+both[0], True))
                    else:
                        control[0] += 1
                    control[1] += 1

                col.append(fill_cell(ws, 'Total', True))
                col.append(fill_cell(ws, 'Total Sistema', True))
                col.append(fill_cell(ws, 'Falta Maxima', True))
                ws.append(col)

            col = []

            # id del obrero
            id = f.find('f36').text
            col.append(fill_cell(ws, int(id)))

            # nombre del obrero
            nombre = f.find('f37').text
            col.append(fill_cell(ws, nombre))

            # Constante de numero de columnas
            n_dias = control[1] - control[0]

            # Agregamos inasistencias y cambiamos formato
            for inasistencia in inasistencias:
                lista = f.find(inasistencia).text
                try:
                    if ord(lista) == 189:
                        lista = 0.5
                    elif ord(lista) == 120:
                        lista = 1
                except:
                    pass
                finally:
                    control[2] += 1
                    if n_dias >= control[2]:
                        col.append(fill_cell(ws, lista))

            # Agregamos la formula para la columna desde 3 hasta n_dias+2
            col.append(fill_cell(ws, create_formula(3, n_dias+2, control[3], control[3], get_column_letter(n_dias+5)
                                                    +str(3))))

            # Agregamos el total
            total = f.find('f53').text
            col.append(fill_cell(ws, int(total)))
            ws.append(col)


        ws[get_column_letter(n_dias+5)+str(3)] = '=MAX('+get_column_letter(n_dias+4)+str(3)+':'+\
                                                 get_column_letter(n_dias+4)+str(control[3])+')'

        ws[get_column_letter(n_dias+5)+str(3)].style = Style(font=Font(name='Calibri', size=11, bold=False),
                           border=Border(left=Side(border_style=borders.BORDER_THIN,color='FF000000'),
                           right=Side(border_style=borders.BORDER_THIN,color='FF000000'),
                           top=Side(border_style=borders.BORDER_THIN, color='FF000000'),
                           bottom=Side(border_style=borders.BORDER_THIN,color='FF000000')))

        ws[get_column_letter(n_dias+5)+str(3)].alignment = Alignment(horizontal='center', vertical='center')

        col = []

        col.append(fill_cell(ws, ''))
        col.append(fill_cell(ws, 'Total Dia', True))

        for idx in range(3, n_dias+3):
            try:
                if calendario[actualDates[idx-3]] is 'libre':
                    col.append(fill_cell(ws, int(0)))
            except KeyError:
                col.append(fill_cell(ws, create_formula(idx, idx, 3, control[3], control[3]-2)))
            except IndexError:
                col.append(fill_cell(ws, create_formula(idx, idx, 3, control[3], control[3]-2)))
        ws.append(col)

        col = []

        '''
        col.append(fill_cell(ws, ''))
        col.append(fill_cell(ws, 'Produccion', True))

        produccion = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 97.14, 0, 21.95,46.45, 0, 0, 70.12, 77.20, 81.92, 82.93, 75.09, 54.33, 0, 54.33]

        for pdx in produccion:
            col.append(fill_cell(ws,pdx ))

        ws.append(col)
        '''
        ws.column_dimensions.group(get_column_letter(1), get_column_letter(n_dias+2), hidden=True)

        params = dict()
        params['use'] = 'bars'

        params['data_min_col'] = n_dias+3
        params['data_max_col'] = n_dias+3
        params['data_min_row'] = 2
        params['data_max_row'] = control[3]

        params['cats_min_col'] = 1
        params['cats_max_col'] = 1
        params['cats_min_row'] = 3
        params['cats_max_row'] = control[3]

        params['type'] = 'bar'
        params['title'] = 'Asistencia por ID'
        params['x_axis'] = 'Operario ID'
        params['y_axis'] = 'Asistencias'
        params['heigth'] = 15
        params['width'] = 30

        # Asistencia Mensual por ID
        mensualID = wb.create_sheet(title="Asistencia por ID")
        ronava_bar_chart(mensualID, ws, params)

        # Asistencia Mensual por Nombre
        params = dict()
        params['use'] = 'bars'
        params['data_min_col'] = n_dias+3
        params['data_max_col'] = n_dias+3
        params['data_min_row'] = 2
        params['data_max_row'] = control[3]

        params['cats_min_col'] = 2
        params['cats_max_col'] = 2
        params['cats_min_row'] = 3
        params['cats_max_row'] = control[3]

        params['type'] = 'bar'
        params['title'] = 'Asistencia por Nombre'
        params['x_axis'] = 'Nombre Operario'
        params['y_axis'] = 'Asistencias'
        params['heigth'] = 15
        params['width'] = 30

        mensualNombre = wb.create_sheet(title="Asistencia por Nombre")
        ronava_bar_chart(mensualNombre, ws, params)

        # Asistencia por dia
        params = dict()
        params['use'] = 'single'
        params['data_min_col'] = 2
        params['data_max_col'] = n_dias + 2
        params['data_min_row'] = control[3]+1
        params['data_max_row'] = control[3]+1

        params['cats_min_col'] = 3
        params['cats_max_col'] = n_dias + 2
        params['cats_min_row'] = 2
        params['cats_max_row'] = 2

        params['type'] = 'col'
        params['title'] = 'Asistencia por dia'
        params['x_axis'] = 'Dia'
        params['y_axis'] = 'Asistencia'
        params['heigth'] = 15
        params['width'] = 40

        mensualDia = wb.create_sheet(title="Asistencia por dia")
        ronava_bar_chart(mensualDia, ws, params)

        '''
        # Asistencia por dia
        params = dict()
        params['use'] = 'prueba'
        params['data_min_col'] = 2
        params['data_max_col'] = n_dias + 2
        params['data_min_row'] = control[3]+1
        params['data_max_row'] = control[3]+1

        params['cats_min_col'] = 3
        params['cats_max_col'] = n_dias + 2
        params['cats_min_row'] = 2
        params['cats_max_row'] = 2

        params['type'] = 'col'
        params['title'] = 'Produccion vs Asistencia'
        params['x_axis'] = 'Dia'
        params['y_axis'] = 'Asistencia'
        params['heigth'] = 15
        params['width'] = 40

        mensualDiaProduccion = wb.create_sheet(title="Produccion vs Asistencia")
        ronava_bar_chart(mensualDiaProduccion, ws, params)

        '''

    wb.save((directorio+"%s"+(simple[:-4]+'.xlsx')) % '\\\\')


def salida(opcion):
    if opcion is not None:
        pass
    else:
        e.msgbox("Cerrando aplicacion")
        sys.exit(0)

yes = True

while yes:
    # TODO change reboot options to spanish
    mensaje = e.msgbox("Transformacion de archivos xml"+"\n"+"\tVersion 1.0.0"+"\n", image='images\\LogoRonava.png')
    salida(mensaje)

    directorio = e.diropenbox(title = "\t\tEscoger directorio", msg = "\n"+"\tSeleccione el directorio con los archivos csv")
    salida(directorio)

    opciones = next(os.walk(directorio))[2]
    salida(opciones)

    # Seleccionamos solo las opciones .csv

    xml = list()

    for file in opciones:
        if file[-4:] == ".xml":
            xml.append(file)

    archivos = e.multchoicebox(msg='Seleccione los archivos a transformar', title='Seleccion de archivos', choices=xml)
    salida(archivos)

    xlsx = list()

    for file in opciones:
        if file[-5:] == ".xlsx":
            xlsx.append(file)

    paraDict = e.choicebox(msg='Seleccionar libres', title='Seleccione el archivo de libres', choices=xlsx)
    salida(paraDict)

    wb = load_workbook((directorio+"%s"+paraDict) % '\\\\')
    ws = wb.active

    actual = 1
    while ws['A'+str(actual)].value is not None:
        calendario[ws['A'+str(actual)].value] = 'libre'
        actual += 1

    salida(e.msgbox("Iniciar transformacion"))

    for file in archivos:
        try:
            transform((directorio+"%s"+file) % '\\\\', file, directorio)
        except IOError:
            msg = "El archivo "+file+" esta abierto en Excel."+"\n"+"Cierrelo para poder transformarlo"
            title = "Error transformando"

            opcion = e.ccbox(msg, title)
            salida(opcion)

    msg = "Reiniciar proceso?"
    title = "Por favor seleccione una opcion"
    if e.ccbox(msg, title):     # show a Continue/Cancel dialog
        pass  # user chose Continue
    else:
        sys.exit(0)     # user chose Cancel



